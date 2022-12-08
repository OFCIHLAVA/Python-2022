import openpyxl as excel
import datetime


# Najde sloupec, kde jsou commenty (Hleda pak v nem info o pridani do ceniku)
def sfe_find_collumn_by_name(pr_nac_sheet, name=str):
    # print(f'V souboru je {pr_nac_sheet.max_collumn} sloupcu.')
    for collumn in range(1, pr_nac_sheet.max_column +1):
        collumn_name = pr_nac_sheet.cell(1, collumn).value
        # print(f'{collumn}. sloupec = {collumn_name}')
        if str(collumn_name).upper() ==name.upper():
            print(f'\n{name} se bude hledat ve sloupci {collumn} = {collumn_name}\n')
            return collumn
# Najde posledni a predposledni pridany radek
def sfe_find_last_and_predposledni_added_row(
    pr_nac_sheet,
    comment_collumn_nr=int,
    pn_collumn_nr=int,
    date_collumn_nr=int,
    ):

    print(f'\nV souboru je {pr_nac_sheet.max_row} radek.\n')
    
    last_two_added_rows = [None, None]
    last_added_item = f'Ještě se nic nepřidávalo'
    last_added_item = f'Ještě se nic nepřidávalo'

    for row in range(2,min(1048576, pr_nac_sheet.max_row+1)):
        comment_cell_to_check = pr_nac_sheet.cell(row, comment_collumn_nr)
        comment = comment_cell_to_check.value
        # if comment != None:
        #     print(f'Řádka {row}, comment: {comment}')

        if "PŘIDÁNO DO CENÍKU" in str(comment).upper():
            last_two_added_rows.pop()
            last_two_added_rows.insert(0, row)
            
            last_added_row, predposledni_added_row = last_two_added_rows[0], last_two_added_rows[-1]
            last_added_date = pr_nac_sheet.cell(row, date_collumn_nr).value
            last_added_item = pr_nac_sheet.cell(row, pn_collumn_nr).value

            print(f'Pridavano do radky: {row}, Item: {last_added_item}, do datumu: {last_added_date}')
    return last_added_row, predposledni_added_row
# Vybere nove nacenene polozky od posledniho pridavani
def sfe_newly_nacenene_lines(pr_nac_sheet,
    posledni_pridana_radka=int,
    comment_column_nr=int,
    pn_collumn_nr=int,
    unit_column_nr=int,
    services_price_column_nr=int,
    ):

    services_linky_pridat_do_ceniku = list()
    linky_nepridavat_comment = list()
    linky_mozna_pridat_comment = list()

    # Tuty commenty pridat, pokud nemaji v sobe taky text z nepridavat_keywords
    pridavat_keywords = (
        "MIX",
        "field service",
        "Galley PN, SN and MSN provided by the customer is a must."
        )
    # Tuty commenty nepridavat.
    nepridavat_keywords = (
        "nedávat do ceníku",
        "nepřidávat do ceníku",
        "aft komplex",
        "nedávat",
        "nepřidávat",
        "lavatory",
        "lav"
        )

    # print(pr_nac_sheet.max_row)
    for row in range(posledni_pridana_radka+1,min(1048576, pr_nac_sheet.max_row+1)):
        # print(row)
        pn = pr_nac_sheet.cell(row, pn_collumn_nr).value
        # print(pn_collumn_nr, pn)
        comment = pr_nac_sheet.cell(row, comment_column_nr).value
        unit = str(pr_nac_sheet.cell(row, unit_column_nr).value)
        platnost_od = datetime.date.today()+datetime.timedelta(2)
        services_cena = pr_nac_sheet.cell(row, services_price_column_nr).value

        if pn != None:
            # Pokud jsou bez commetu, rovnou pridat.
            if comment == None:
                services_linky_pridat_do_ceniku.append(["PPB200018","", str(pn).replace(" ",""),"USD", unit, 0, "Not Applicable", platnost_od, services_cena, unit]) 
            # Pokud Maji nejaky comment
            else:
                # Nepridavane commenty rovnou nepridavat
                for text in nepridavat_keywords:
                    if text.lower() in str(comment).lower():
                        linky_nepridavat_comment.append(row)
                        break                                                    
                # Pokud se nebrejklo vyse (neni to nepridavaci comment), da se do nacenovanych linek, pokud se comment == textu v pridavat keywords, jinak se da do mozna pridal linek.
                else:
                    for text in pridavat_keywords:
                        if text.lower() == str(comment).lower():
                            services_linky_pridat_do_ceniku.append(["PPB200018","", str(pn).replace(" ",""),"USD", unit, 0, "Not Applicable", platnost_od, services_cena, unit]) 
                            break     
                    else:
                        # Pridat v formatu [[linka, comment], [services linka]] aby se pozdeji daly pripadne snadno pridat k uz hotovym linkam. 
                        linky_mozna_pridat_comment.append([
                        [row, comment],
                        ["PPB200018","", str(pn).replace(" ",""),"USD", unit, 0, "Not Applicable", platnost_od, services_cena, unit]
                        ])
        else:
            print(f'Vsechny linky projety.')
            posledni_nacenena_linka_v_souboru = row-1
            break
    return services_linky_pridat_do_ceniku, linky_nepridavat_comment, linky_mozna_pridat_comment, posledni_nacenena_linka_v_souboru

# Nacteni EXCEL souboru. (Otevreni v rezimu only values - necte vzorce ze souboru, jen hodnoty.)
print("1. Otevirani souboru prubezne aktualizace . . .")

sfe_prubezna_aktualizace = excel.load_workbook("Y:\\Departments\\Sales and Marketing\\Aftersales\\02_FRONT OFFICE\\03_PRICING\\2022\\SFE\\SFE 2022 prubezna aktualizace.xlsx", data_only=True)
spa = sfe_prubezna_aktualizace
sheet_pr_nac = spa["Prubezne nacenovani"]
print("OK\n")

# Nalezeni potrebnych sloupcu.
print("2. Hledani cisel sloupcu v souboru . . .")

comment_column = sfe_find_collumn_by_name(sheet_pr_nac, "comment")
pn_column = sfe_find_collumn_by_name(sheet_pr_nac, "P/N")
date_column = sfe_find_collumn_by_name(sheet_pr_nac, "date of entry")
uom_column = sfe_find_collumn_by_name(sheet_pr_nac, "unit")
usd_services_column = sfe_find_collumn_by_name(sheet_pr_nac, "2022 Safran Services USD (15 percent discount)")

print("OK\n")

# Nalezeni posledni pridane linky v prubeznem nacenovani.
print("3. Hledani posledni pridane radky do pr. nac. . . .")

posledni_a_prespodleni_pridana_radka = sfe_find_last_and_predposledni_added_row(sheet_pr_nac, comment_column, pn_column, date_column)
posledni_pridana_radka = posledni_a_prespodleni_pridana_radka[0]
predposledni_pridana_radka = posledni_a_prespodleni_pridana_radka[-1]

print("OK\n")
print(f'\nPosledni pridana radka {posledni_pridana_radka} a predposledni pridana radka {predposledni_pridana_radka}.\n')

# Ziskani seznamu urcite pridavanych, nepridavanych a mozna pridavanych linek.
print("4. Ziskani seznamu polozek pridat do LN ceniku . . .")

services_linky, nepridavane_linky, zkontrolovat_jestli_pridavat_linky, posledni_nacenena_linka_v_souboru = sfe_newly_nacenene_lines(sheet_pr_nac, posledni_pridana_radka, comment_column, pn_column, uom_column, usd_services_column)

print("OK\n")

### Closnuti SFE nacenovaci tabulky, aby se pak dala znovu otervrit i se vzorci a upravit commenty.
spa.close()

print(f'Pridavano {len(services_linky)} SFE linek do Services ceniku:\n')
# print(f'EUR linky:\n')
# for line in services_linky:
#     print(line)
# print()
# print(f'USD linky:\n')
# for line in usd_linky:
#     print(line)

print(f'Celkem {len(nepridavane_linky)} linek nepridava ceniku na zaklade komentu z prubezneho nacenovani:\n')
print(f'Nepridavane linky z pr. nac. cislo:\n')
for line in nepridavane_linky:
    print(line)

print(f'\nCelkem {len(zkontrolovat_jestli_pridavat_linky)} linek nevim, jestli mam pridat do ceniku (maji nejaky comment):')
print(f'Nevim jestli pridavat linky:\n')
for line in zkontrolovat_jestli_pridavat_linky:
    # line[0][0] = cislo linky, line[0][1] = comment linky
    print(f'Linka: {line[0][0]}, Comment: {line[0][1]}')

# Moznost rucne dopsat, ktere linky jeste pridat.
if len(zkontrolovat_jestli_pridavat_linky) >0:
    dodatecne_pridat_linky = input(f'\nJestli chces nektere z techto linek vyse take pridat do ceniku, zadej cisla jejich linek (u vice linek oddel carkou) a stiskni ENTER . . .\nJinak jen stiskni ENTER pro pokracovani.\n')
    dodatecne_pridat_linky = [linka.replace(" ","") for linka in dodatecne_pridat_linky.split(",")]

    print(f'Pridam tedy jeste linky:\n')
    for cislo_linky in dodatecne_pridat_linky:
        for linka in zkontrolovat_jestli_pridavat_linky:
            # print(linka)
            # print(cislo_linky, type(cislo_linky))
            # print(linka[0][0], type(linka[0][0]))

            if cislo_linky == str(linka[0][0]):
                print(f'Linku: {cislo_linky} dodatecne pridavam do ceniku - OK.\n')
                # linka[1] = services nacenena linka
                services_linky.append(linka[1])
                # Pokud se presune do pridat do ceniku, odsud odebrat
                zkontrolovat_jestli_pridavat_linky.remove(linka)
                break
        else:
            print(f'Tuhle linku - {cislo_linky} - jsem v seznamu nenasel, preskakuji . . .\n')

    # Ty co jsme nechteli dodatecne pridat se presunou do seznamu nepridavanych linek.
    for line in zkontrolovat_jestli_pridavat_linky:
        nepridavane_linky.append(int(line[0][0]))
    nepridavane_linky = sorted(set(nepridavane_linky))

### Zapsani vysledku do USD xls souboru

# USD
# Otevreni USD souboru
print("5. Zapisovni do USD souboru vysktupu . . .")

usd_excel = excel.load_workbook("Y:\\Departments\\Sales and Marketing\\Aftersales\\11_PLANNING\\23_Python_utilities\\9_LN_ceníky\\měsíční přidávání\\SFE\\USD SFE Pricelist PB200018.xlsx") 
usd_sheet1 = usd_excel.worksheets[0]

# Smazani starych dat
usd_sheet1.delete_rows(3,usd_sheet1.max_row)

# Zapsani linek do souboru
for i, linka in enumerate(services_linky):
    for u, udaj in enumerate(linka):
        cell_to_update = usd_sheet1.cell(3+i, 4+u)
        cell_to_update.value = udaj
usd_excel.save("Y:\\Departments\\Sales and Marketing\\Aftersales\\11_PLANNING\\23_Python_utilities\\9_LN_ceníky\\měsíční přidávání\\SFE\\USD SFE Pricelist PB200018.xlsx")
usd_excel.close()
print("OK\n")

# Tisk vysledku.
print(f'Hotovo\n')
print(f'Nakonec tedy \n')

print(f'Pridavano {len(services_linky)} SFE linek do Services ceniku:\n')
print(f'Services linky:\n')
for line in services_linky:
    print(line)
print()

print(f'\nCelkem {len(nepridavane_linky)} linek nepridavam ceniku na zaklade komentu z prubezneho nacenovani:\n')
print(f'Nepridavane linky z pr. nac. cislo:\n')
for line in nepridavane_linky:
    print(line)

# Jako posledni vec zapsat do pr. nac. souboru pridano do ceniku s dnesnim datumem do COMMENTU a ulozit soubor
print("6. Zapisovni commentu o pridani do ceniku do pr. nac. souboru . . .")

sfe_prubezna_aktualizace = excel.load_workbook("Y:\\Departments\\Sales and Marketing\\Aftersales\\02_FRONT OFFICE\\03_PRICING\\2022\\SFE\\SFE 2022 prubezna aktualizace.xlsx")
spa = sfe_prubezna_aktualizace
sheet_pr_nac = spa["Prubezne nacenovani"]

cell_pridano_do_ceniku = sheet_pr_nac.cell(posledni_nacenena_linka_v_souboru, comment_column)
puvodni_comment = str(cell_pridano_do_ceniku.value)

if puvodni_comment == "None":
    puvodni_comment = ""    
novy_comment = f'přidáno do ceníku{" " + puvodni_comment}'

cell_pridano_do_ceniku.value = novy_comment

sfe_prubezna_aktualizace.save("Y:\\Departments\\Sales and Marketing\\Aftersales\\02_FRONT OFFICE\\03_PRICING\\2022\\SFE\\SFE 2022 prubezna aktualizace.xlsx")
sfe_prubezna_aktualizace.close()
print("OK\n")

print(f'\nHOTOVO.\n')
input(f'\nUkonci program stiskem klavesy ENTER . . . ')