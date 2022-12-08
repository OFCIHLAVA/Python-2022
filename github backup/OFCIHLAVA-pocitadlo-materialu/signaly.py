from pathlib import Path
import webbrowser
import openpyxl as xl
import os

# cesta do slozky, kde jsou soubory ke zpracovani
# test path
# path = Path('Y:\\Departments\\Sales and Marketing\\Aftersales\\11_PLANNING\\16_Aftersales_master_plan\\Archiv')
# glob metoda vrati vsechny soubory v dane slozce jako generator objekt --> lze v nem iterovat

def signal(searchString, inColumn, filename, outputFile):
    file_to_process = xl.load_workbook(filename)
    output_file = xl.load_workbook(outputFile)
    input_sheet = file_to_process.worksheets[0]
    output_sheet = output_file.worksheets[0]
    # Zkopiruje zahlavi, pokud tam jeste neni.
    if output_sheet.cell(1,1).value == None:
        # print(1)
        for column in range(1, input_sheet.max_column +1):
            headings_to_copy = input_sheet.cell(1, column).value
            output_sheet.cell(1, column).value = headings_to_copy
    # Pokud bude treba najit sloupec se "Signalem" nejprve:
    # for column in range(1,input_sheet.max_column +1):
    #     singal_cell_check = input_sheet.cell(1,column)
    #     if type(signal_cell_to_check) == str and "signal".upper() in singal_cell_check.value.upper():
    #         column_to_check = column
    #         break
    # Pro vsechny radky ve zpracovavanem souboru.
    if input_sheet.cell(1,7).value.upper() != "SIGNAL":
        #print("2a")
        print(f'Sloupec {inColumn} v souboru {filename} nema oznaceni signal, ale {input_sheet.cell(1,7).value}.')
    else:
        # print("2b")
        # print(input_sheet.max_row)
        for row in range(1,min(1048576, input_sheet.max_row+1)):
            # Kontrola sedmeho sloupce jestli obsahuje hledany retezec
            cell_to_check = input_sheet.cell(row+1,inColumn).value        
            if cell_to_check != None and type(cell_to_check) == str and searchString.upper() in cell_to_check.upper():
                print(f'Ano {searchString} na radce {row+1} v souboru {filename}')
                row_to_write = output_sheet.max_row +1
                for column in range(1, input_sheet.max_column +1):
                    cell_to_copy = input_sheet.cell(row+1, column).value
                    output_sheet.cell(row_to_write, column).value = cell_to_copy
    file_to_process.close()
    output_file.save(f'{path}\\{output_file_name}')           

# Samotny program:

print(f'Program projde vsechny excelovske soubory v zadane slozce a se zadanou priponou a vyhleda zadany retezec v zadanem sloupci. Vysledek ulozi do souboru "output.xls" v prohledavane slozce.\n')

valid_path = False
while not valid_path:
    path = Path(input(f'Zadej cestu do slozky: ').replace("\\","\\\\"))
    if os.path.exists(path):
        valid_path = True
    else:
        print(f'POZOR! Neplatna cesta do slozky. Zkus zadat jinou cestu do slozky se soubory ke zpracovani.')

valid_extension = False
while not valid_extension:
    extension = input(f'Zadej priponu typu souboru, ktere chces projit ve forme ".xxx": ')
    if type(extension) == str and extension[0] == "." and extension.count(".") == 1:
        valid_extension = True
    else:
        print(f'Neplatna pripona. Pripona musi byt text zacinajici teckou ".xxx"')

string_to_search_for = (input(f'Jaky retezec chces hledat? (nerozlisuje velka a mala pismena): '))

is_int = False
while not is_int:
    try:
        column_to_search = int(input(f'V kolikatem sloupci chces heldat? (zadej cislo sloupce - A = 1, B = 2 ...): '))
        if type(column_to_search) == int:
            is_int = True
    except ValueError:
        print(f'Je potreba zadat cislo sloupce.')

wb = xl.Workbook()
output_file_name = 'output.xlsx'    
ws1 = wb.active
ws1.title = "output"
wb.save(f'{path}\\{output_file_name}')

files_to_process = []
for file_path in path.glob(f'*{extension}'):
    file = str(file_path).split('\\')[-1]
    files_to_process.append(file)
for file in files_to_process:
    if "$" in file:
        files_to_process.remove(file)
print(f'Pocet souboru ke zpracovani: {len(files_to_process)}\n')
# print(files_to_process)

already_done_count = 0
for file in files_to_process:
    print(file)
    signal(string_to_search_for,column_to_search,f'{path}\\{file}',f'{path}\\{output_file_name}')
    already_done_count +=1
    print(f'Hotovo {already_done_count}/{len(files_to_process)}')

print(f'Konec programu. Vysledek ulozen do souboru {output_file_name} ve slozce {str(path)}')

# Udelat kontrolu pro ty soubory, kde je to nejak divne posunute v master planu
